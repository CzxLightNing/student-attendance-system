"""
签到考勤应用的视图定义
Attendance app view definitions
包含学生签到、签到历史、教师签到码管理、考勤详情、导出 Excel 等功能
Includes student check-in, history, teacher code management, attendance details, Excel export, etc.
所有视图优先使用 Django Class-Based Views 实现
All views use Django Class-Based Views by preference
"""
import json
import random
import logging
from datetime import datetime, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.views.generic import View, TemplateView, ListView, DetailView
from django.views.decorators.http import require_POST
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, FileResponse, HttpResponse
from django.contrib import messages
from django.utils import timezone
from django.db.models import Count, Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from accounts.decorators import StudentRequiredMixin, TeacherRequiredMixin
from accounts.models import CustomUser
from .models import Class, AttendanceCode, AttendanceRecord

# 创建审计日志记录器
# Create audit logger
audit_logger = logging.getLogger('audit')


# ==================== 学生签到视图 ====================
# ==================== Student Check-in Views ====================

class StudentHomeView(LoginRequiredMixin, StudentRequiredMixin, TemplateView):
    """
    学生主页视图
    Student home view
    显示签到码输入框，学生可输入 4 位签到码进行签到
    Display code input; student enters 4-digit code to check in
    """
    template_name = 'student/index.html'

    def get_context_data(self, **kwargs):
        """获取当前学生所在班级的有效签到码信息"""
        """Get active code info for the current student's class"""
        context = super().get_context_data(**kwargs)
        student = self.request.user

        if student.student_class:
            # 查找该学生班级当前有效的签到码
            # Find active code for this student's class
            active_code = AttendanceCode.objects.filter(
                class_group=student.student_class,
                is_active=True,
                expires_at__gt=timezone.now()
            ).order_by('-created_at').first()

            if active_code:
                # 检查该学生是否已经针对此签到码签到
                # Check whether the student has already used this code
                already_signed = AttendanceRecord.objects.filter(
                    student=student,
                    attendance_code=active_code
                ).exists()
                context['active_code'] = active_code
                context['already_signed'] = already_signed

        return context


class StudentCheckinView(LoginRequiredMixin, StudentRequiredMixin, View):
    """
    学生签到 AJAX 接口（POST，返回 JSON）
    校验签到码 → 校验班级归属 → 防重复签到 → 记录签到
    Validate code → verify class → check duplicates → record check-in
    """

    def post(self, request):
        """
        处理学生签到请求
        返回 JSON 格式的签到结果
        """
        student = request.user

        # 检查学生是否已分配班级
        # Check if student has been assigned to a class
        if not student.student_class:
            return JsonResponse({
                'success': False,
                'message': '您尚未被分配到任何班级，请联系管理员'
            }, status=400)

        try:
            # 解析请求体中的 JSON 数据
            # Parse JSON data from request body
            data = json.loads(request.body)
            code_input = data.get('code', '').strip()
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': '请求数据格式错误'
            }, status=400)

        # 校验签到码是否为 4 位数字
        # Validate code is a 4-digit number
        if not code_input or len(code_input) != 4 or not code_input.isdigit():
            return JsonResponse({
                'success': False,
                'message': '请输入 4 位数字签到码'
            }, status=400)

        # 查找有效的签到码
        # Find active check-in code
        active_code = AttendanceCode.objects.filter(
            class_group=student.student_class,
            code=code_input,
            is_active=True,
            expires_at__gt=timezone.now()
        ).order_by('-created_at').first()

        if not active_code:
            return JsonResponse({
                'success': False,
                'message': '签到码无效或已过期，请确认后重试'
            }, status=400)

        # 校验学生是否属于签到码对应的班级
        # Verify student belongs to the code's class
        if student.student_class != active_code.class_group:
            return JsonResponse({
                'success': False,
                'message': '该签到码不属于您所在的班级'
            }, status=400)

        # 检查是否重复签到
        # Check for duplicate sign-in
        if AttendanceRecord.objects.filter(student=student, attendance_code=active_code).exists():
            return JsonResponse({
                'success': False,
                'message': '您已经签到过了，无需重复签到'
            }, status=400)

        # 判断签到状态：正常 或 迟到
        # Determine check-in status: Normal or Late
        now = timezone.now()
        status = '正常'
        if active_code.course_start_time:
            # 如果设置了课程开始时间，签到时间晚于课程时间 则标记为迟到
            # If course start time is set and check-in time is later, mark as late
            if now > active_code.course_start_time + timedelta(minutes=5):
                status = '迟到'

        # 创建签到记录
        # Create attendance record
        record = AttendanceRecord.objects.create(
            student=student,
            attendance_code=active_code,
            status=status
        )

        audit_logger.info(f'学生签到 - 用户名:{student.username} 班级:{student.student_class.name} 签到码:{code_input} 状态:{status}')

        return JsonResponse({
            'success': True,
            'message': f'签到成功！状态：{status}',
            'data': {
                'timestamp': record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'status': record.status,
            }
        })


class StudentHistoryView(LoginRequiredMixin, StudentRequiredMixin, ListView):
    """
    学生签到历史视图
    分页展示当前学生的签到记录列表，支持按日期范围筛选
    """
    template_name = 'student/history.html'
    context_object_name = 'records'
    paginate_by = 15

    def get_queryset(self):
        """筛选当前学生的签到记录，支持日期范围过滤"""
        """Filter current student check-in records, support date range filter"""
        queryset = AttendanceRecord.objects.filter(
            student=self.request.user
        ).select_related('attendance_code__class_group').order_by('-timestamp')

        # 日期范围筛选
        # Date range filtering
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')

        if date_from:
            queryset = queryset.filter(timestamp__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(timestamp__date__lte=date_to)

        return queryset

    def get_context_data(self, **kwargs):
        """向模板传递筛选参数，用于表单回填"""
        """Pass filter params to template for form backfill"""
        context = super().get_context_data(**kwargs)
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        return context


# ==================== 教师管理视图 ====================
# ==================== Teacher Management Views ====================

class TeacherHomeView(LoginRequiredMixin, TeacherRequiredMixin, ListView):
    """
    教师主页视图
    列出当前教师管理的所有班级列表
    List all classes managed by current teacher
    """
    template_name = 'teacher/index.html'
    context_object_name = 'classes'

    def get_queryset(self):
        """获取当前教师管理的班级，按年级和名称排序"""
        """Get classes managed by current teacher, sorted by grade and name"""
        return self.request.user.managed_classes.all().order_by('grade', 'name')


class TeacherClassDetailView(LoginRequiredMixin, TeacherRequiredMixin, DetailView):
    """
    班级签到管理视图
    教师可在此页面：生成签到码、查看当前有效签到码（含倒计时）、
    查看已签到/未签到学生区域（实时轮询更新）
    """
    template_name = 'teacher/class_detail.html'
    model = Class
    context_object_name = 'class_group'
    pk_url_kwarg = 'class_id'

    def get_queryset(self):
        """仅允许教师访问自己管理的班级"""
        return self.request.user.managed_classes.all()

    def get_context_data(self, **kwargs):
        """向模板传递当前班级的有效签到码、签到统计等信息"""
        """Pass active code, stats, and other info for current class to template"""
        context = super().get_context_data(**kwargs)
        class_group = self.object

        # 查找当前班级的有效签到码
        # Find active code for current class
        active_code = AttendanceCode.objects.filter(
            class_group=class_group,
            is_active=True,
            expires_at__gt=timezone.now()
        ).order_by('-created_at').first()

        context['active_code'] = active_code

        # 获取已签到学生信息
        # Get signed-in student info
        if active_code:
            signed_records = AttendanceRecord.objects.filter(
                attendance_code=active_code
            ).select_related('student')
            signed_students = [r.student for r in signed_records]
            context['signed_students'] = signed_students
            context['signed_count'] = len(signed_students)

            # 获取未签到学生列表
            # Get unsigned student list
            all_students = list(CustomUser.objects.filter(
                role='student',
                student_class=class_group
            ))
            context['unsigned_students'] = [
                s for s in all_students if s not in signed_students
            ]
            context['total_students'] = len(all_students)
        else:
            # 获取该班级所有学生
            # Get all students in this class
            all_students = CustomUser.objects.filter(
                role='student',
                student_class=class_group
            )
            context['all_students'] = all_students
            context['total_students'] = all_students.count()

        return context


class TeacherGenerateCodeView(LoginRequiredMixin, TeacherRequiredMixin, View):
    """
    生成签到码 AJAX 接口（POST，返回 JSON）
    random.randint(1000, 9999)，设置 expires_at 和可选的 course_start_time
    新码生成时自动将旧码失效
    """

    def post(self, request, class_id):
        """处理生成签到码请求"""
        teacher = request.user
        class_group = get_object_or_404(Class, id=class_id)

        # 验证教师是否有权管理该班级
        # Verify teacher has permission to manage this class
        if not teacher.managed_classes.filter(id=class_id).exists():
            return JsonResponse({
                'success': False,
                'message': '您无权管理该班级'
            }, status=403)

        try:
            data = json.loads(request.body)
            duration = int(data.get('duration', 10))  # 默认 10 分钟
            course_start = data.get('course_start', None)  # 可选课程开始时间
        except (json.JSONDecodeError, ValueError):
            duration = 10
            course_start = None

        # 限制有效时长范围：1 ~ 120 分钟
        # Limit duration range: 1 ~ 120 minutes
        if duration < 1 or duration > 120:
            duration = 10

        # 生成新签到码前，失效该班级所有旧签到码
        # Deactivate all old codes for this class before generating new one
        AttendanceCode.objects.filter(
            class_group=class_group,
            is_active=True
        ).update(is_active=False)

        # 生成 4 位随机数字签到码
        # Generate 4-digit random check-in code
        code = str(random.randint(1000, 9999))
        now = timezone.now()
        expires_at = now + timedelta(minutes=duration)

        # 解析课程开始时间（如果提供了）
        # Parse course start time (if provided)
        course_start_time = None
        if course_start:
            try:
                course_start_time = datetime.fromisoformat(course_start)
                if timezone.is_naive(course_start_time):
                    course_start_time = timezone.make_aware(course_start_time)
            except (ValueError, TypeError):
                course_start_time = now

        # 创建签到码
        # Create check-in code
        attendance_code = AttendanceCode.objects.create(
            class_group=class_group,
            code=code,
            created_by=teacher,
            expires_at=expires_at,
            course_start_time=course_start_time,
            is_active=True,
        )

        audit_logger.info(f'生成签到码 - 班级:{class_group.name} 签到码:{code} 有效期:{duration}分钟 教师:{teacher.username}')

        return JsonResponse({
            'success': True,
            'message': f'签到码生成成功：{code}',
            'data': {
                'id': attendance_code.id,
                'code': code,
                'expires_at': expires_at.isoformat(),
                'created_at': attendance_code.created_at.isoformat(),
                'duration': duration,
            }
        })


class TeacherDeactivateCodeView(LoginRequiredMixin, TeacherRequiredMixin, View):
    """
    失效签到码 AJAX 接口（POST）
    教师手动失效当前有效的签到码
    Teacher manually deactivates current active code
    """

    def post(self, request, class_id):
        """处理失效签到码请求"""
        teacher = request.user
        class_group = get_object_or_404(Class, id=class_id)

        # 验证教师权限
        # Verify teacher permission
        if not teacher.managed_classes.filter(id=class_id).exists():
            return JsonResponse({
                'success': False,
                'message': '您无权管理该班级'
            }, status=403)

        # 失效该班级所有有效签到码
        # Deactivate all active codes for this class
        updated_count = AttendanceCode.objects.filter(
            class_group=class_group,
            is_active=True
        ).update(is_active=False)

        audit_logger.info(f'手动失效签到码 - 班级:{class_group.name} 教师:{teacher.username}')

        if updated_count > 0:
            return JsonResponse({'success': True, 'message': '签到码已失效'})
        else:
            return JsonResponse({'success': False, 'message': '当前没有有效的签到码'})


class TeacherCheckinStatusView(LoginRequiredMixin, TeacherRequiredMixin, View):
    """
    签到实时状态查询 AJAX 接口（GET，返回 JSON）
    返回该签到码下已签到学生列表和未签到学生列表
    Return signed and unsigned student lists for this code
    供前端 setInterval 轮询使用
    For use with frontend setInterval polling
    """

    def get(self, request, code_id):
        """查询指定签到码的实时签到状态"""
        """Query real-time status of specified check-in code"""
        attendance_code = get_object_or_404(AttendanceCode, id=code_id)

        # 验证教师权限
        # Verify teacher permission
        if not request.user.managed_classes.filter(id=attendance_code.class_group_id).exists():
            return JsonResponse({
                'success': False,
                'message': '您无权查看该班级'
            }, status=403)

        # 获取已签到学生
        # Get signed-in students
        signed_records = AttendanceRecord.objects.filter(
            attendance_code=attendance_code
        ).select_related('student').order_by('-timestamp')

        signed_students_list = []
        for record in signed_records:
            signed_students_list.append({
                'id': record.student.id,
                'name': f"{record.student.last_name}{record.student.first_name}",
                'student_id': record.student.student_id or '',
                'timestamp': record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'status': record.status,
            })

        # 获取班级所有学生
        # Get all students in class
        all_students = CustomUser.objects.filter(
            role='student',
            student_class=attendance_code.class_group
        )
        signed_student_ids = set(r.student_id for r in signed_records)

        unsigned_students_list = []
        for student in all_students:
            if student.id not in signed_student_ids:
                unsigned_students_list.append({
                    'id': student.id,
                    'name': f"{student.last_name}{student.first_name}",
                    'student_id': student.student_id or '',
                })

        return JsonResponse({
            'success': True,
            'data': {
                'code': attendance_code.code,
                'is_active': attendance_code.is_active,
                'expires_at': attendance_code.expires_at.isoformat(),
                'signed_count': len(signed_students_list),
                'total_count': all_students.count(),
                'signed_students': signed_students_list,
                'unsigned_students': unsigned_students_list,
            }
        })


class TeacherAttendanceDetailView(LoginRequiredMixin, TeacherRequiredMixin, DetailView):
    """
    签到码考勤详情视图
    展示某次签到码对应的完整签到记录
    """
    template_name = 'teacher/attendance_detail.html'
    model = AttendanceCode
    context_object_name = 'attendance_code'
    pk_url_kwarg = 'code_id'

    def get_queryset(self):
        """仅允许教师访问自己管理的班级的签到码详情"""
        return AttendanceCode.objects.filter(
            class_group__in=self.request.user.managed_classes.all()
        ).select_related('class_group', 'created_by')

    def get_context_data(self, **kwargs):
        """获取该签到码下已签到和未签到学生的完整信息"""
        """Get complete signed/unsigned student info for this code"""
        context = super().get_context_data(**kwargs)
        attendance_code = self.object

        # 获取签到记录
        # Get check-in records
        records = AttendanceRecord.objects.filter(
            attendance_code=attendance_code
        ).select_related('student').order_by('-timestamp')

        context['records'] = records
        context['signed_count'] = records.count()

        # 获取班级所有学生
        # Get all students in class
        all_students = CustomUser.objects.filter(
            role='student',
            student_class=attendance_code.class_group
        )
        signed_student_ids = set(r.student_id for r in records)

        context['unsigned_students'] = [
            s for s in all_students if s.id not in signed_student_ids
        ]
        context['total_students'] = all_students.count()

        return context


class TeacherExportExcelView(LoginRequiredMixin, TeacherRequiredMixin, View):
    """
    导出考勤记录到 Excel
    使用 openpyxl 生成 .xlsx 文件并返回 FileResponse
    """

    def get(self, request, code_id):
        """导出指定签到码的签到记录为 Excel 文件"""
        """Export check-in records for specified code as Excel file"""
        attendance_code = get_object_or_404(AttendanceCode, id=code_id)

        # 验证教师权限
        # Verify teacher permission
        if not request.user.managed_classes.filter(id=attendance_code.class_group_id).exists():
            messages.error(request, '您无权导出该班级的考勤数据')
            return redirect('attendance:teacher_home')

        # 获取签到记录
        # Get check-in records
        records = AttendanceRecord.objects.filter(
            attendance_code=attendance_code
        ).select_related('student').order_by('-timestamp')

        # 获取未签到学生
        # Get unsigned students
        all_students = CustomUser.objects.filter(
            role='student',
            student_class=attendance_code.class_group
        )
        signed_student_ids = set(r.student_id for r in records)
        unsigned_students = [s for s in all_students if s.id not in signed_student_ids]

        # 创建 Excel 工作簿
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '考勤记录'

        # 定义样式
        # Define styles
        header_font = Font(name='微软雅黑', bold=True, size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        # 写入标题行
        # Write title row
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f'考勤签到记录 - {attendance_code.class_group.name}（签到码：{attendance_code.code}）'
        title_cell.font = Font(name='微软雅黑', bold=True, size=14)
        title_cell.alignment = center_alignment

        # 写入基本信息
        # Write basic info
        ws.merge_cells('A2:F2')
        ws['A2'].value = f'创建时间：{attendance_code.created_at.strftime("%Y-%m-%d %H:%M:%S")}    教师：{attendance_code.created_by.last_name}{attendance_code.created_by.first_name}'
        ws['A2'].font = Font(name='微软雅黑', size=10)

        # 写入表头
        # Write headers
        headers = ['序号', '学号', '姓名', '签到时间', '签到状态', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # 写入签到数据
        # Write check-in data
        row = 5
        for index, record in enumerate(records, 1):
            student = record.student
            ws.cell(row=row, column=1, value=index).alignment = center_alignment
            ws.cell(row=row, column=2, value=student.student_id or '-').alignment = center_alignment
            ws.cell(row=row, column=3, value=f"{student.last_name}{student.first_name}")
            ws.cell(row=row, column=4, value=record.timestamp.strftime('%Y-%m-%d %H:%M:%S')).alignment = center_alignment
            ws.cell(row=row, column=5, value=record.status).alignment = center_alignment
            ws.cell(row=row, column=6, value='')
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

        # 写入缺勤学生
        # Write absent students
        for student in unsigned_students:
            ws.cell(row=row, column=1, value=row - 4).alignment = center_alignment
            ws.cell(row=row, column=2, value=student.student_id or '-').alignment = center_alignment
            ws.cell(row=row, column=3, value=f"{student.last_name}{student.first_name}")
            ws.cell(row=row, column=4, value='-').alignment = center_alignment
            ws.cell(row=row, column=5, value='缺勤').alignment = center_alignment
            ws.cell(row=row, column=6, value='')
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

        # 写入统计信息
        # Write statistics
        row += 1
        ws.cell(row=row, column=1, value=f'统计：应到 {all_students.count()} 人，实到 {records.count()} 人，缺勤 {len(unsigned_students)} 人')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True)

        # 设置列宽
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15

        # 生成文件名
        # Generate filename
        filename = f"考勤_{attendance_code.class_group.name}_{attendance_code.code}_{attendance_code.created_at.strftime('%Y%m%d%H%M')}.xlsx"

        # 创建 HTTP 响应
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)

        audit_logger.info(f'导出考勤Excel - 班级:{attendance_code.class_group.name} 签到码:{attendance_code.code} 教师:{request.user.username}')

        return response


class TeacherAttendanceHistoryView(LoginRequiredMixin, TeacherRequiredMixin, ListView):
    """
    教师查看班级考勤历史记录
    展示所选班级历次签到码及其签到统计
    """
    template_name = 'teacher/attendance_history.html'
    context_object_name = 'codes'
    paginate_by = 10

    def get_queryset(self):
        """获取教师管理的指定班级的签到码列表"""
        """Get check-in code list for specified class managed by teacher"""
        class_id = self.kwargs.get('class_id')
        return AttendanceCode.objects.filter(
            class_group_id=class_id,
            class_group__in=self.request.user.managed_classes.all()
        ).select_related('class_group').order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        class_id = self.kwargs.get('class_id')
        context['class_group'] = get_object_or_404(
            Class,
            id=class_id,
            id__in=self.request.user.managed_classes.values_list('id', flat=True)
        )
        return context


class TeacherStudentListView(LoginRequiredMixin, TeacherRequiredMixin, DetailView):
    """
    教师查看班级学生名单
    显示所管理班级的完整学生列表（学号、姓名等）
    """
    template_name = 'teacher/student_list.html'
    model = Class
    context_object_name = 'class_group'
    pk_url_kwarg = 'class_id'

    def get_queryset(self):
        """仅允许教师访问自己管理的班级"""
        return self.request.user.managed_classes.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['students'] = CustomUser.objects.filter(
            role='student',
            student_class=self.object
        ).order_by('student_id')
        return context
